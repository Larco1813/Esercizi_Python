import tkinter as tk
from openpyxl import Workbook, load_workbook
from tkcalendar import DateEntry

filename = "C:\\Users\\gabri\\OneDrive\\Desktop\\Python\\esercizi\\FinanzeGui\\dati_soldi.xlsx"

try:
    workbook = load_workbook(filename)
    sheet = workbook.active
except FileNotFoundError:
    if filename.path.exists(filename):
        filename.remove(filename)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Data", "Importo"])


root = tk.Tk()


root.title("Gestisci Finanze")
root.geometry("350x500") 
data_label = tk.Label(root, text="Inserisci la data (formato YYYY/MM/DD):")
data_label.pack()
data_entry = DateEntry(root, width=12, background='darkblue', foreground='white', borderwidth=2)
data_entry.pack(padx=10, pady=10)
importo_label = tk.Label(root, text="Inserisci l'importo:")
importo_label.pack()
importo_entry = tk.Entry(root)
importo_entry.pack()






def inserisci_dati():
    data = data_entry.get()
    importo = importo_entry.get()
    if importo:
        sheet.append([data, importo])
        workbook.save(filename)
        print("Dati salvati con successo nel file Excel.")
    else:
        print("Errore: Inserire un importo valido.")

def cancella_ultima_riga():
    max_row = sheet.max_row
    if max_row > 1: 
        sheet.delete_rows(max_row)
        workbook.save(filename)
        print("Ultima riga cancellata con successo.")
    else:
        print("Errore: Nessuna riga da cancellare.")


button_inserisci = tk.Button(root, text="Aggiungi", command=inserisci_dati)
button_inserisci.pack()
button_cancella = tk.Button(root, text="Cancella Ultima Riga", command=cancella_ultima_riga)
button_cancella.pack()

root.mainloop()
